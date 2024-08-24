from flask import Flask, render_template, request
from openpyxl import load_workbook, Workbook
import os

app = Flask(__name__)

# Initialize the Excel file if it doesn't exist
def initialize_excel_file(filename):
    if not os.path.exists(filename):
        workbook = Workbook()
        sheet = workbook.active
        # Add headers if needed
        sheet.append(["Year", "Month", "Day", "Time", "Event", "Size", "Post Food", "Timestamp"])
        workbook.save(filename)

@app.route('/', methods=['GET', 'POST'])
def index():
    initialize_excel_file("Kopiko_Habit_Tracker.xlsx")

    # Initialize response variables
    summary = ""
    last_entries = []
    timestamps = []

    if request.method == 'POST':
        if 'delete' in request.form:
            # Handle delete request
            timestamp_to_delete = request.form['timestamp']
            workbook = load_workbook("Kopiko_Habit_Tracker.xlsx")
            sheet = workbook.active

            # Find and delete the row with the matching timestamp
            for row in sheet.iter_rows(min_row=2, max_col=8):
                if row[7].value == timestamp_to_delete:
                    sheet.delete_rows(row[0].row)
                    break

            workbook.save("Kopiko_Habit_Tracker.xlsx")
            summary = "Entry deleted successfully."

        elif 'submit' in request.form:
            # Handle submit request
            year = request.form['year']
            month = request.form['month']
            day = request.form['day']
            time = request.form['time']
            event = request.form['event']
            size = request.form['size']
            post_food = request.form['post_food']
            timestamp = f"{year}-{month[:3]}-{day} {time}"
            
            # Load the existing Excel workbook
            workbook = load_workbook("Kopiko_Habit_Tracker.xlsx")
            sheet = workbook.active

            # Find the next available row in the sheet
            next_row = sheet.max_row + 1

            # Write the data to the next row
            sheet.cell(row=next_row, column=1, value=year)
            sheet.cell(row=next_row, column=2, value=month)
            sheet.cell(row=next_row, column=3, value=day)
            sheet.cell(row=next_row, column=4, value=time)
            sheet.cell(row=next_row, column=5, value=event)
            sheet.cell(row=next_row, column=6, value=size)
            sheet.cell(row=next_row, column=7, value=post_food)
            sheet.cell(row=next_row, column=8, value=timestamp)

            # Save the workbook
            workbook.save("Kopiko_Habit_Tracker.xlsx")
            summary = "Entry added successfully."

        # Load the workbook again to read the last 10 entries and timestamps
        workbook = load_workbook("Kopiko_Habit_Tracker.xlsx")
        sheet = workbook.active

        # Get the last 10 entries
        num_rows = sheet.max_row
        last_10_entries = []
        for row in range(max(num_rows - 9, 2), num_rows + 1):  # Start from row 2 to avoid header
            entry = [sheet.cell(row=row, column=col).value for col in range(1, 9)]
            last_10_entries.append(entry)

        # Get all timestamps for deletion
        timestamps = [sheet.cell(row=row, column=8).value for row in range(2, num_rows + 1)]

    else:
        # Handle initial GET request
        workbook = load_workbook("Kopiko_Habit_Tracker.xlsx")
        sheet = workbook.active
        timestamps = [sheet.cell(row=row, column=8).value for row in range(2, sheet.max_row + 1)]

    return render_template('index.html', summary=summary, last_entries=last_entries, timestamps=timestamps)

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
