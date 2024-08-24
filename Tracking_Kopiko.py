import tkinter as tk
from tkinter import ttk
from openpyxl import load_workbook
from datetime import datetime

# Function to handle the submission of data
def submit_data():
    year = year_entry.get()
    month = month_entry.get()
    day= day_entry.get()
    time = time_entry.get()
    event = event_var.get()
    size = size_var.get()
    post_food = post_food_var.get()

    # Load the existing Excel workbook
    workbook = load_workbook("Kopiko_Habit_Tracker.xlsx")
    sheet = workbook.active

    # Find the next available row in the sheet
    next_row = sheet.max_row + 1

    # Write the data to the next row
    sheet.cell(row=next_row, column=1, value=year)
    sheet.cell(row=next_row, column=2, value=month)
    sheet.cell(row=next_row, column=3, value=day)
    sheet.cell(row=next_row, column=4, value=time)
    sheet.cell(row=next_row, column=5, value=event)
    sheet.cell(row=next_row, column=6, value=size)
    sheet.cell(row=next_row, column=7, value=post_food)

    # Save the workbook
    workbook.save("Kopiko_Habit_Tracker.xlsx")

    # Update the summary label with the submitted data
    summary_label.config(text=f"Successfully submitted: Event: {event} at {time}, {day}/{month}/{year}, Size: {size}")

    # Clear the form after submission
    # year_entry.set("")
    # month_entry.set("")
    # day_entry.set("")
    # time_entry.delete(0, tk.END)
    # event_var.set("")
    # size_var.set("")
    # post_food_var.set("")

# Create the main window
root = tk.Tk()
root.title("Kopiko Habit Tracker")

# Create and place the labels and entry fields
tk.Label(root, text="Year").grid(row=1, column=0, padx=10, pady=5)
year_entry = tk.StringVar()
event_combobox = ttk.Combobox(root, textvariable=year_entry)
event_combobox['values'] = (2024, 2025)
event_combobox.grid(row=1, column=1, padx=10, pady=5)

tk.Label(root, text="Month").grid(row=2, column=0, padx=10, pady=5)
month_entry = tk.StringVar()
event_combobox = ttk.Combobox(root, textvariable=month_entry)
event_combobox['values'] = ("January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December")
event_combobox.grid(row=2, column=1, padx=10, pady=5)

tk.Label(root, text="Day").grid(row=3, column=0, padx=10, pady=5)
day_entry = tk.StringVar()
event_combobox = ttk.Combobox(root, textvariable=day_entry)
event_combobox['values'] = (1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24,25,26,27,28,29,30,31)
event_combobox.grid(row=3, column=1, padx=10, pady=5)

tk.Label(root, text="Time - 24 hr (HH:MM):").grid(row=4, column=0, padx=10, pady=5)
time_entry = tk.Entry(root)
time_entry.grid(row=4, column=1, padx=10, pady=5)

tk.Label(root, text="Event:").grid(row=5, column=0, padx=10, pady=5)
event_var = tk.StringVar()
event_combobox = ttk.Combobox(root, textvariable=event_var)
event_combobox['values'] = ("Eat", "Poop")
event_combobox.grid(row=5, column=1, padx=10, pady=5)

tk.Label(root, text="Size:").grid(row=6, column=0, padx=10, pady=5)
size_var = tk.StringVar()
size_combobox = ttk.Combobox(root, textvariable=size_var)
size_combobox['values'] = ("Small poop / Half cup", "Medium / Full cup", "Large / More than 1 cup")
size_combobox.grid(row=6, column=1, padx=10, pady=5)

tk.Label(root, text="After Food?:").grid(row=7, column=0, padx=10, pady=5)
post_food_var = tk.StringVar()
size_combobox = ttk.Combobox(root, textvariable=post_food_var)
size_combobox['values'] = (1, 0)
size_combobox.grid(row=7, column=1, padx=10, pady=5)

# Create and place the submit button
submit_button = tk.Button(root, text="Submit", command=submit_data)
submit_button.grid(row=8, column=0, columnspan=2, padx=10, pady=20)

summary_label = tk.Label(root, text="")
summary_label.grid(row=9, column=0, columnspan=2, padx=10, pady=10)

# Run the GUI event loop
root.mainloop()